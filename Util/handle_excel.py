# coding=utf-8
import logging
import os

import openpyxl

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s -  %(message)s')


class HandExcel:
    def __init__(self, index):
        self.index = index      # excel的sheet数
        # 获取当前路径
        self.base_path = os.path.dirname(os.path.abspath(__file__))
        self.parent_dir = os.path.dirname(self.base_path)
        self.excelroute = r'\case\test_case.xlsx'

    def load_excel(self):
        """
        加载excel
        :excelroute: excel文件记录
        """
        open_excel = openpyxl.load_workbook(self.parent_dir + self.excelroute)
        return open_excel

    def get_sheet_data(self):
        """
        加载所有的sheet的内容
        """
        sheet_name = self.load_excel().sheetnames
        if self.index is None:
            self.index = 0
        data = self.load_excel()[sheet_name[self.index]]
        return data

    def get_cell_value(self, row, cols, index=None):
        """
        获取单元格的内容
        :row: 行数
        :cols: 列数
        :index: sheet数
        """
        cell_value = self.get_sheet_data().cell(row=row, column=cols).value
        return cell_value

    def get_rows(self):
        """
        获取行数
        """
        row = self.get_sheet_data().max_row
        return row

    def get_rows_value(self, row):
        """
        获取当前sheet中的行内容
        :row: 获取的行数
        """

        row_list = []
        row_first_row = []
        row_dict = {}
        # 获取测试用例集
        for rows in self.get_sheet_data()[row]:
            row_list.append(rows.value)

        # 获取测试用例集标题
        for rows_first in self.get_sheet_data()[1]:
            row_first_row.append(rows_first.value)

        # 将测试用例标题和测试用例组合
        for rows in range(len(row_first_row)):
            row_dict[row_first_row[rows]] = row_list[rows]

        return row_dict


execl_data = HandExcel

# if __name__ == '__main__':
#     handle = HandExcel(0)
#     print(handle.get_rows_value(2))
